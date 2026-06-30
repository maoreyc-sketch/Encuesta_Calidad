import streamlit as st
import pandas as pd
import csv
import os
import io
import base64
import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side)
from openpyxl.utils import get_column_letter

# =====================================================================
# CONFIGURACIÓN VISUAL
# =====================================================================
st.set_page_config(page_title="Encuesta Calidad MEN", layout="wide")

st.markdown("""
    <style>
        .block-container { max-width: 1100px; padding-top: 2rem; padding-bottom: 2rem; }
        .stButton>button { width: 100%; }
        div[data-testid="metric-container"] {
            background: #f8f4ff; border: 1px solid #e0d4f5;
            border-radius: 10px; padding: 12px 18px;
        }
        div[data-testid="metric-container"] label { color: #4E1F79 !important; font-weight: 600; }
    </style>
""", unsafe_allow_html=True)

# =====================================================================
# ENCABEZADO INSTITUCIONAL
# =====================================================================
col_logo, col_tit = st.columns([3, 4])
with col_logo:
    st.image("logo_min.jpeg", width=520)
with col_tit:
    st.markdown("""
        <div style='text-align: left;'>
            <h1 style='font-family: Arial; font-size: 20pt; font-weight: bold; margin-bottom: 0;'>
                Ministerio de Educación Nacional
            </h1>
            <h2 style='font-family: Arial; font-size: 16pt; font-weight: bold; color: #4E1F79;'>
                Indice Integral de Calidad Educativa
            </h2>
        </div>
    """, unsafe_allow_html=True)
st.markdown("<hr>", unsafe_allow_html=True)

# =====================================================================
# ALMACENAMIENTO: GITHUB (persistente) + CSV LOCAL (respaldo)
# Configurar en Streamlit > Settings > Secrets:
#   github_token = "ghp_xxxxxxxxxxxx"
#   github_repo  = "usuario/nombre-repositorio"
# =====================================================================
CSV_PATH      = 'Consolidado_Encuestas_MEN.csv'
GITHUB_FILE   = 'Consolidado_Encuestas_MEN.csv'
_gh_disponible = (
    "github_token" in st.secrets and
    "github_repo"  in st.secrets
)

def _gh_headers():
    tk = st.secrets["github_token"]
    # ghp_ = clasico (token), github_pat_ = fine-grained (Bearer)
    prefijo = "Bearer" if tk.startswith("github_pat_") else "token"
    return {
        "Authorization": f"{prefijo} {tk}",
        "Accept": "application/vnd.github.v3+json",
        "X-GitHub-Api-Version": "2022-11-28"
    }

def _gh_url():
    repo   = st.secrets["github_repo"]
    branch = st.secrets.get("github_branch", "main")
    return f"https://api.github.com/repos/{repo}/contents/{GITHUB_FILE}", branch

def cargar_desde_github():
    """Lee el CSV desde GitHub usando la URL Raw (sin límites de tamaño) y obtiene el SHA de forma ligera."""
    try:
        url_api, branch = _gh_url()
        repo = st.secrets["github_repo"]
        
        # 1. Descargar el contenido usando la URL RAW (Soporta archivos gigantes sin límite)
        url_raw = f"https://raw.githubusercontent.com/{repo}/{branch}/{GITHUB_FILE}"
        r_raw = requests.get(url_raw, headers=_gh_headers(), timeout=10)
        
        # 2. Consultar la API solo para traer el código SHA de control
        r_api = requests.get(url_api, headers=_gh_headers(), params={"ref": branch}, timeout=10)
        sha = r_api.json().get("sha") if r_api.status_code == 200 else None

        if r_raw.status_code == 200:
            contenido = r_raw.content.decode("utf-8-sig")
            df = pd.read_csv(io.StringIO(contenido), engine='python',on_bad_lines='skip') # Déjalo por seguridad, pero el motor 'python' corregirá el parseo)
            return df, sha
        return pd.DataFrame(), None
    except Exception as e:
        # Imprime el error interno en la consola del codespace por si necesitas debuguear
        print(f"Error en cargar_desde_github: {str(e)}")
        return pd.DataFrame(), None

def guardar_en_github(df_completo):
    """
    Sube el DataFrame completo al repositorio GitHub de manera robusta.
    """
    try:
        url, branch = _gh_url()
        
        # SOLUCIÓN: Obtener el SHA de forma ligera sin descargar el contenido completo
        sha = None
        headers_sha = _gh_headers()
        r_sha = requests.get(url, headers=headers_sha, params={"ref": branch}, timeout=10)
        if r_sha.status_code == 200:
            sha = r_sha.json().get("sha")

        contenido_csv = df_completo.to_csv(index=False, encoding='utf-8-sig', quoting=csv.QUOTE_ALL)
        contenido_b64 = base64.b64encode(contenido_csv.encode('utf-8-sig')).decode()
        
        payload = {
            "message": f"Encuesta registrada {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            "content": contenido_b64,
            "branch": branch
        }
        if sha:
            payload["sha"] = sha
            
        r = requests.put(url, json=payload, headers=_gh_headers(), timeout=15)
        return r.status_code in (200, 201), r.text
    except Exception as e:
        return False, str(e)
# ── Lectura unificada (GitHub primero, CSV local de respaldo) ─────────
def leer_datos():
    if _gh_disponible:
        df, _ = cargar_desde_github()
        if not df.empty:
            return df
    # Respaldo: CSV local
    if os.path.exists(CSV_PATH):
        try:
            return pd.read_csv(CSV_PATH, encoding='utf-8-sig', on_bad_lines='skip')
        except:
            pass
    return pd.DataFrame()

def es_duplicado(email):
    df = leer_datos()
    if df.empty or 'EMAIL_VALIDADO' not in df.columns:
        return False
    return email in df['EMAIL_VALIDADO'].str.strip().str.lower().values

def es_duplicado_dane(codigo_dane):
    """Verifica duplicado por Código DANE (más robusto que por email)."""
    if not codigo_dane:
        return False
    df = leer_datos()
    if df.empty or 'DANE' not in df.columns:
        return False
    return codigo_dane in df['DANE'].astype(str).str.strip().values

def guardar_encuesta(final):
    """
    1. Lee todos los registros actuales.
    2. Agrega la fila nueva.
    3. Guarda en GitHub (persistente) Y en CSV local (respaldo).
    """
    try:
        df_actual = leer_datos()
        df_nuevo  = pd.DataFrame([final])
        df_total  = pd.concat([df_actual, df_nuevo], ignore_index=True) if not df_actual.empty else df_nuevo

        errores = []

        # — Guardar en GitHub (persistente) —
        if _gh_disponible:
            ok_gh, err_gh = guardar_en_github(df_total)
            if not ok_gh:
                errores.append(f"GitHub: {err_gh}")

        # — Guardar CSV local (siempre, como respaldo) —
        encabezado = not os.path.exists(CSV_PATH)
        df_nuevo.to_csv(
            CSV_PATH, mode='a', header=encabezado,
            index=False, encoding='utf-8-sig', quoting=csv.QUOTE_ALL
        )

        # Si GitHub falló pero CSV local funcionó, avisamos pero no bloqueamos
        if errores and _gh_disponible:
            return True, f"⚠️ Guardado localmente. Error en GitHub: {'; '.join(errores)}"
        return True, ''

    except Exception as e:
        return False, str(e)

# =====================================================================
# GENERADOR DE EXCEL PROFESIONAL
# =====================================================================
def generar_excel(df, df_maestra):
    """
    Genera un Excel con 4 hojas:
      1. Respuestas Completas  (datos crudos formateados)
      2. Resumen por Departamento
      3. Establecimientos Pendientes
      4. Estadísticas Generales
    """
    wb = Workbook()

    # ── Estilos reutilizables ────────────────────────────────────────
    MORADO      = "4E1F79"
    MORADO_CLARO= "E8D9F5"
    VERDE       = "D4EDDA"
    AMARILLO    = "FFF3CD"
    ROJO        = "F8D7DA"
    BLANCO      = "FFFFFF"
    GRIS        = "F2F2F2"

    fill_header  = PatternFill("solid", fgColor=MORADO)
    fill_sub     = PatternFill("solid", fgColor=MORADO_CLARO)
    fill_alt     = PatternFill("solid", fgColor=GRIS)
    font_header  = Font(bold=True, color=BLANCO, size=11)
    font_titulo  = Font(bold=True, color=MORADO, size=13)
    font_normal  = Font(size=10)
    font_bold    = Font(bold=True, size=10)
    al_center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    al_left      = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    borde = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    def estilo_fila(ws, fila, num_cols, par=True):
        fill = fill_alt if par else PatternFill("solid", fgColor=BLANCO)
        for col in range(1, num_cols + 1):
            c = ws.cell(row=fila, column=col)
            c.fill   = fill
            c.border = borde
            c.font   = font_normal
            c.alignment = al_left

    def auto_ancho(ws, min_w=12, max_w=50):
        for col in ws.columns:
            ancho = max(
                (len(str(c.value)) if c.value else 0) for c in col
            )
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(ancho + 3, min_w), max_w)

    # ──────────────────────────────────────────────────────────────────
    # HOJA 1: RESPUESTAS COMPLETAS
    # ──────────────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Respuestas Completas"

    # Título
    ws1.merge_cells("A1:F1")
    t = ws1["A1"]
    t.value     = f"Consolidado Encuesta Calidad Educativa — Ministerio de Educación Nacional"
    t.font      = font_titulo
    t.alignment = al_center
    t.fill      = fill_sub

    ws1.merge_cells("A2:F2")
    f = ws1["A2"]
    f.value     = f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}   |   Total registros: {len(df)}"
    f.font      = Font(italic=True, size=9, color="666666")
    f.alignment = al_center

    ws1.row_dimensions[1].height = 22
    ws1.row_dimensions[2].height = 16

    if not df.empty:
        cols = list(df.columns)
        # Encabezados
        for ci, col in enumerate(cols, start=1):
            c = ws1.cell(row=3, column=ci, value=col)
            c.fill      = fill_header
            c.font      = font_header
            c.alignment = al_center
            c.border    = borde
        ws1.row_dimensions[3].height = 20
        ws1.freeze_panes = "A4"

        # Datos
        for ri, row in enumerate(df.itertuples(index=False), start=4):
            for ci, val in enumerate(row, start=1):
                c = ws1.cell(row=ri, column=ci, value=val)
                c.border    = borde
                c.font      = font_normal
                c.alignment = al_left
            if ri % 2 == 0:
                for ci in range(1, len(cols)+1):
                    ws1.cell(row=ri, column=ci).fill = fill_alt

        # Destacar columna FECHA
        if 'FECHA' in cols:
            idx_fecha = cols.index('FECHA') + 1
            ws1.column_dimensions[get_column_letter(idx_fecha)].width = 20

    auto_ancho(ws1)

    # ──────────────────────────────────────────────────────────────────
    # HOJA 2: RESUMEN POR DEPARTAMENTO
    # ──────────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumen por Departamento")

    ws2.merge_cells("A1:G1")
    t2 = ws2["A1"]
    t2.value     = "Avance de Encuestas por Departamento"
    t2.font      = font_titulo
    t2.alignment = al_center
    t2.fill      = fill_sub
    ws2.row_dimensions[1].height = 22

    cabeceras = ["Departamento", "Esperadas", "Recibidas", "Pendientes", "% Avance", "Última Respuesta", "Estado"]
    for ci, cab in enumerate(cabeceras, 1):
        c = ws2.cell(row=2, column=ci, value=cab)
        c.fill      = fill_header
        c.font      = font_header
        c.alignment = al_center
        c.border    = borde
    ws2.freeze_panes = "A3"

    if not df.empty and not df_maestra.empty:
        col_dep_m = 'DEPARTAMENTO' if 'DEPARTAMENTO' in df_maestra.columns else None
        col_dep_r = 'DEPTO'        if 'DEPTO'        in df.columns         else None

        if col_dep_m and col_dep_r:
            esp = df_maestra.groupby(col_dep_m).size().reset_index(name='Esperadas')
            rec = df.groupby(col_dep_r).size().reset_index(name='Recibidas')
            rec.rename(columns={col_dep_r: col_dep_m}, inplace=True)
            ult = df.groupby(col_dep_r)['FECHA'].max().reset_index()
            ult.rename(columns={col_dep_r: col_dep_m, 'FECHA': 'Ultima'}, inplace=True)
            avance = esp.merge(rec, on=col_dep_m, how='left').merge(ult, on=col_dep_m, how='left').fillna(0)
            avance['Recibidas'] = avance['Recibidas'].astype(int)
            avance['Pendientes'] = avance['Esperadas'] - avance['Recibidas']
            avance['Pct'] = (avance['Recibidas'] / avance['Esperadas'] * 100).round(1)
            avance = avance.sort_values('Pct', ascending=False).reset_index(drop=True)

            for fila_num, (_, row) in enumerate(avance.iterrows(), start=3):
                pct = float(row['Pct'])
                estado = "✅ Completo" if pct >= 100 else ("🟡 En progreso" if pct >= 50 else "🔴 Crítico")
                fill_estado = (PatternFill("solid", fgColor=VERDE) if pct >= 100
                               else (PatternFill("solid", fgColor=AMARILLO) if pct >= 50
                               else PatternFill("solid", fgColor=ROJO)))
                vals = [row[col_dep_m], int(row['Esperadas']), int(row['Recibidas']),
                        int(row['Pendientes']), f"{pct}%", row.get('Ultima', ''), estado]
                for ci, v in enumerate(vals, 1):
                    c = ws2.cell(row=fila_num, column=ci, value=v)
                    c.border    = borde
                    c.alignment = al_center if ci > 1 else al_left
                    c.font      = font_bold if ci == 5 else font_normal
                    if ci == 7:
                        c.fill = fill_estado

    auto_ancho(ws2)

    # ──────────────────────────────────────────────────────────────────
    # HOJA 3: ESTABLECIMIENTOS PENDIENTES
    # ──────────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Establecimientos Pendientes")

    ws3.merge_cells("A1:F1")
    t3 = ws3["A1"]
    t3.value     = "Establecimientos que NO han respondido la encuesta"
    t3.font      = font_titulo
    t3.alignment = al_center
    t3.fill      = PatternFill("solid", fgColor=ROJO)
    ws3.row_dimensions[1].height = 22

    cols_pend = ['NOMBRE_ESTABLECIMIENTO', 'DEPARTAMENTO', 'SECRETARIA', 'RECTOR', 'EMAIL', 'TELEFONO']

    if not df_maestra.empty and 'EMAIL' in df_maestra.columns:
        emails_ok = (
            set(df['EMAIL_VALIDADO'].str.strip().str.lower().values)
            if not df.empty and 'EMAIL_VALIDADO' in df.columns else set()
        )
        pendientes = df_maestra[~df_maestra['EMAIL'].isin(emails_ok)].copy()
        cols_disp  = [c for c in cols_pend if c in pendientes.columns]

        ws3.merge_cells(f"A2:{get_column_letter(len(cols_disp))}2")
        info = ws3["A2"]
        info.value     = f"Total pendientes: {len(pendientes)}   |   Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        info.font      = Font(italic=True, size=9, color="721C24")
        info.alignment = al_center

        for ci, col in enumerate(cols_disp, 1):
            c = ws3.cell(row=3, column=ci, value=col)
            c.fill      = fill_header
            c.font      = font_header
            c.alignment = al_center
            c.border    = borde
        ws3.freeze_panes = "A4"

        for ri, row in enumerate(pendientes[cols_disp].itertuples(index=False), start=4):
            for ci, val in enumerate(row, 1):
                c = ws3.cell(row=ri, column=ci, value=val)
                c.border    = borde
                c.font      = font_normal
                c.alignment = al_left
            if ri % 2 == 0:
                for ci in range(1, len(cols_disp)+1):
                    ws3.cell(row=ri, column=ci).fill = fill_alt

    auto_ancho(ws3)

    # ──────────────────────────────────────────────────────────────────
    # HOJA 4: ESTADÍSTICAS GENERALES
    # ──────────────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Estadísticas Generales")

    ws4.merge_cells("A1:C1")
    t4 = ws4["A1"]
    t4.value     = "Estadísticas Generales — Encuesta Calidad MEN"
    t4.font      = font_titulo
    t4.alignment = al_center
    t4.fill      = fill_sub
    ws4.row_dimensions[1].height = 22
    ws4.column_dimensions['A'].width = 30
    ws4.column_dimensions['B'].width = 20
    ws4.column_dimensions['C'].width = 20

    total_esp  = len(df_maestra) if not df_maestra.empty else 0
    total_rec  = len(df)         if not df.empty         else 0
    total_pend = max(total_esp - total_rec, 0)
    pct_avance = round(total_rec / total_esp * 100, 1) if total_esp > 0 else 0

    stats = [
        ("INDICADOR",                       "VALOR",        ""),
        ("Total establecimientos esperados", total_esp,      ""),
        ("Encuestas recibidas",              total_rec,      ""),
        ("Establecimientos pendientes",      total_pend,     ""),
        ("% de avance general",             f"{pct_avance}%",""),
        ("",                                 "",             ""),
        ("Fecha de corte",   datetime.now().strftime('%d/%m/%Y'), ""),
        ("Hora de corte",    datetime.now().strftime('%H:%M:%S'), ""),
    ]

    if not df.empty and 'DEPTO' in df.columns:
        stats.append(("Departamentos con respuestas", df['DEPTO'].nunique(), ""))

    for ri, (a, b, c_val) in enumerate(stats, start=2):
        ca = ws4.cell(row=ri, column=1, value=a)
        cb = ws4.cell(row=ri, column=2, value=b)
        cc = ws4.cell(row=ri, column=3, value=c_val)
        if ri == 2:  # encabezado
            for cell in [ca, cb, cc]:
                cell.fill = fill_header; cell.font = font_header; cell.alignment = al_center
        else:
            ca.font = font_bold;  ca.alignment = al_left
            cb.font = font_normal; cb.alignment = al_center
            if ri % 2 == 0:
                for cell in [ca, cb]: cell.fill = fill_alt
        for cell in [ca, cb, cc]:
            cell.border = borde

    # ── Serializar a bytes ───────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# =====================================================================
# FUNCIONES AUXILIARES
# =====================================================================
def scroll_al_inicio():
    js = "<script>window.parent.document.querySelector('.main').scrollTo(0,0);</script>"
    st.components.v1.html(js, height=0)

@st.cache_data
def cargar_maestra():
    try:
        df = pd.read_excel("EE 2026.xlsx", sheet_name=0, dtype=str)
        if 'EMAIL' in df.columns:
            df['EMAIL'] = df['EMAIL'].str.strip().str.lower()
        return df
    except Exception as e:
        st.error(f"Error cargando EE 2026.xlsx: {e}")
        return pd.DataFrame()

# =====================================================================
# ESTADO DE MEMORIA
# =====================================================================
hojas = ["Infraestructura", "Computadores", "Recursos Pedagógicos", "Programas", "Docentes"]

if 'iniciado'          not in st.session_state: st.session_state.iniciado          = False
if 'encuesta_iniciada' not in st.session_state: st.session_state.encuesta_iniciada = False
if 'paso'              not in st.session_state: st.session_state.paso              = 0
if 'editando'          not in st.session_state: st.session_state.editando          = False
if 'finalizado'        not in st.session_state: st.session_state.finalizado        = False
if 'resp_enc'          not in st.session_state: st.session_state.resp_enc          = {}
if 'oblig'             not in st.session_state: st.session_state.oblig             = []
if 'panel_admin'       not in st.session_state: st.session_state.panel_admin       = False
if 'hacer_scroll'      not in st.session_state: st.session_state.hacer_scroll      = False
if 'codigo_dane_sel'   not in st.session_state: st.session_state.codigo_dane_sel   = ''

# =====================================================================
# PANEL DE ADMINISTRACIÓN (pantalla completa)
# =====================================================================
def mostrar_panel_admin():
    st.markdown("## 📊 Panel de Administración — Encuesta Calidad MEN")

    # Indicador de almacenamiento activo
    # ── Diagnóstico REAL de almacenamiento ──────────────────────────
    csv_local_existe = os.path.exists(CSV_PATH)

    if _gh_disponible:
        # Verificar que GitHub realmente responde y tiene el archivo
        df_gh, sha_gh = cargar_desde_github()
        gh_tiene_datos = not df_gh.empty
        if gh_tiene_datos:
            st.success(f"🔒 **GitHub activo y verificado** — {len(df_gh)} registros confirmados en el repositorio.")
        else:
            st.error(
                "⚠️ **GitHub configurado pero SIN datos confirmados.** "
                "Los secrets existen pero la escritura no ha llegado al repositorio. "
                "Los registros actuales están solo en el servidor temporal de Streamlit."
            )
            if csv_local_existe:
                st.info("📁 Se encontró CSV local en el servidor. Usa el botón de rescate para sincronizarlo con GitHub.")
                if st.button("🔄 Rescatar CSV local → subir a GitHub ahora"):
                    df_rescate = pd.read_csv(CSV_PATH, encoding='utf-8-sig', on_bad_lines='skip')
                    ok, err = guardar_en_github(df_rescate)
                    if ok:
                        st.success("✅ ¡Datos sincronizados con GitHub correctamente!")
                        st.rerun()
                    else:
                        st.error(f"Error al subir: {err}. Verifica que el token tenga permisos de escritura (scope: repo).")
    else:
        st.warning("⚠️ GitHub no configurado — datos solo en CSV local (se pierden al reiniciar).")

    st.markdown("---")

    # ── Botón limpiar registros de prueba ───────────────────────────
    with st.expander("🧹 Limpiar registros de prueba (usar antes de publicar)"):
        st.warning("⚠️ Esto borrará TODOS los registros actuales. Úsalo solo para limpiar pruebas antes del lanzamiento real.")
        confirmar = st.checkbox("Confirmo que quiero eliminar todos los registros de prueba")
        if confirmar:
            if st.button("🗑️ Eliminar todos los registros ahora", type="primary"):
                # Borrar CSV local
                if os.path.exists(CSV_PATH):
                    os.remove(CSV_PATH)
                # Borrar de GitHub si existe
                if _gh_disponible:
                    try:
                        url, branch = _gh_url()
                        _, sha_del = cargar_desde_github()
                        if sha_del:
                            requests.delete(url, headers=_gh_headers(), json={
                                "message": "Limpieza registros de prueba",
                                "sha": sha_del,
                                "branch": branch
                            }, timeout=10)
                    except:
                        pass
                st.success("✅ Registros eliminados. La base está limpia para encuestas reales.")
                st.rerun()

    st.markdown("---")

    df         = leer_datos()
    df_maestra = cargar_maestra()

    # ── KPIs ────────────────────────────────────────────────────────
    total_esp  = len(df_maestra) if not df_maestra.empty else 0
    total_rec  = len(df)         if not df.empty         else 0
    pendientes = max(total_esp - total_rec, 0)
    pct        = round(total_rec / total_esp * 100, 1) if total_esp > 0 else 0

    k1, k2, k3, k4 = st.columns(4)
    k1.metric("📋 Total Esperado",     f"{total_esp}")
    k2.metric("✅ Encuestas Recibidas", f"{total_rec}")
    k3.metric("⏳ Pendientes",          f"{pendientes}")
    k4.metric("📈 Avance General",      f"{pct}%")

    st.markdown(f"""
        <div style='background:#e0d4f5;border-radius:10px;height:22px;margin:10px 0 4px 0;'>
            <div style='background:#4E1F79;width:{pct}%;height:22px;border-radius:10px;
                        display:flex;align-items:center;justify-content:center;
                        color:white;font-size:13px;font-weight:bold;'>
                {pct}%
            </div>
        </div>
        <p style='text-align:right;font-size:11px;color:#888;margin:0;'>
            Corte: {datetime.now().strftime('%d/%m/%Y %H:%M')}
        </p>
    """, unsafe_allow_html=True)

    st.markdown("---")

    # ── BOTÓN DE DESCARGA EXCEL (siempre visible arriba) ────────────
    if not df.empty:
        nombre_excel = f"Encuesta_Calidad_MEN_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        excel_bytes  = generar_excel(df, df_maestra)
        st.download_button(
            label="📥 Descargar Base Completa en Excel (.xlsx)",
            data=excel_bytes,
            file_name=nombre_excel,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )
        st.caption("El Excel incluye 4 hojas: Respuestas Completas · Resumen por Departamento · Pendientes · Estadísticas")
        st.markdown("---")
    else:
        st.info("Aún no hay encuestas registradas.")
        if st.button("⬅️ Volver al inicio"):
            st.session_state.panel_admin = False
            st.rerun()
        return

    # ── Tabs de visualización web ────────────────────────────────────
    tab1, tab2, tab3, tab4 = st.tabs([
        "📊 Avance por Secretaria",
        "🏫 Establecimientos Pendientes",
        "📋 Ver Todas las Respuestas",
        "⬇️ Otras Descargas"
    ])

    # ── TAB 1: Avance por Secretaria ────────────────────────────────
    with tab1:
        col_sec_m = 'SECRETARIA' if not df_maestra.empty and 'SECRETARIA' in df_maestra.columns else None
        col_sec_r = 'MUNI'       if 'MUNI' in df.columns else None

        if col_sec_m and col_sec_r:
            esp = df_maestra.groupby(col_sec_m).size().reset_index(name='Esperadas')
            rec = df.groupby(col_sec_r).size().reset_index(name='Recibidas')
            rec.rename(columns={col_sec_r: col_sec_m}, inplace=True)
            avance = esp.merge(rec, on=col_sec_m, how='left').fillna(0)
            avance['Recibidas']  = avance['Recibidas'].astype(int)
            avance['Pendientes'] = avance['Esperadas'] - avance['Recibidas']
            avance['% Avance']   = (avance['Recibidas'] / avance['Esperadas'] * 100).round(1)
            avance = avance.sort_values('% Avance', ascending=False).reset_index(drop=True)
            avance = avance.rename(columns={col_sec_m: 'Secretaría'})

            def colorear_pct(val):
                if isinstance(val, float):
                    if val >= 80:  return 'background-color:#d4edda;color:#155724'
                    if val >= 50:  return 'background-color:#fff3cd;color:#856404'
                    return 'background-color:#f8d7da;color:#721c24'
                return ''

            st.dataframe(avance.style.map(colorear_pct, subset=['% Avance']),
                         use_container_width=True, hide_index=True)
            st.bar_chart(avance.set_index('Secretaría')['% Avance'])
        else:
            st.info("Sin datos suficientes para calcular avance por Secretaría.")

    # ── TAB 2: Pendientes ───────────────────────────────────────────
    with tab2:
        if not df_maestra.empty and 'EMAIL' in df_maestra.columns:
            emails_ok  = (set(df['EMAIL_VALIDADO'].str.strip().str.lower().values)
                          if 'EMAIL_VALIDADO' in df.columns else set())
            pend_df    = df_maestra[~df_maestra['EMAIL'].isin(emails_ok)].copy()
            cols_m     = [c for c in ['NOMBRE_ESTABLECIMIENTO','DEPARTAMENTO','SECRETARIA','RECTOR','EMAIL','TELEFONO']
                          if c in pend_df.columns]

            st.markdown(f"**{len(pend_df)} establecimientos sin responder**")
            if 'DEPARTAMENTO' in pend_df.columns:
                deptos = ['Todos'] + sorted(pend_df['DEPARTAMENTO'].dropna().unique().tolist())
                filtro = st.selectbox("Filtrar por departamento:", deptos, key="f_pend")
                if filtro != 'Todos':
                    pend_df = pend_df[pend_df['DEPARTAMENTO'] == filtro]

            st.dataframe(pend_df[cols_m] if cols_m else pend_df,
                         use_container_width=True, hide_index=True)
        else:
            st.info("No se puede calcular pendientes sin la maestra de establecimientos.")

    # ── TAB 3: Ver respuestas ────────────────────────────────────────
    with tab3:
        fc1, fc2 = st.columns(2)
        with fc1:
            f_dep = 'Todos'
            if 'DEPTO' in df.columns:
                deps = ['Todos'] + sorted(df['DEPTO'].dropna().unique().tolist())
                f_dep = st.selectbox("Filtrar por Departamento:", deps, key="f_dep_r")
        with fc2:
            busq = st.text_input("🔍 Buscar:", key="busq_r")

        df_f = df.copy()
        if f_dep != 'Todos':
            df_f = df_f[df_f['DEPTO'] == f_dep]
        if busq:
            mask = df_f.apply(lambda r: r.astype(str).str.contains(busq, case=False).any(), axis=1)
            df_f = df_f[mask]

        st.dataframe(df_f, use_container_width=True, hide_index=True)
        st.caption(f"Mostrando {len(df_f)} de {len(df)} registros")

    # ── TAB 4: Otras descargas ───────────────────────────────────────
    with tab4:
        dc1, dc2 = st.columns(2)
        with dc1:
            st.markdown("**📄 CSV — Respuestas completas**")
            st.download_button(
                "Descargar CSV",
                data=df.to_csv(index=False).encode('utf-8-sig'),
                file_name=f"Consolidado_MEN_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                mime='text/csv', use_container_width=True
            )
        with dc2:
            if not df_maestra.empty and 'EMAIL' in df_maestra.columns:
                emails_ok  = (set(df['EMAIL_VALIDADO'].str.strip().str.lower().values)
                              if 'EMAIL_VALIDADO' in df.columns else set())
                pend_dl    = df_maestra[~df_maestra['EMAIL'].isin(emails_ok)]
                st.markdown("**⏳ CSV — Solo pendientes**")
                st.download_button(
                    "Descargar pendientes CSV",
                    data=pend_dl.to_csv(index=False).encode('utf-8-sig'),
                    file_name=f"Pendientes_MEN_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                    mime='text/csv', use_container_width=True
                )

    st.markdown("---")
    if st.button("⬅️ Volver al inicio de la encuesta"):
        st.session_state.panel_admin = False
        st.rerun()

# =====================================================================
# FLUJO PRINCIPAL
# =====================================================================
if st.session_state.panel_admin:
    mostrar_panel_admin()
    st.stop()

if st.session_state.finalizado:
    st.success("✅ ¡Encuesta procesada exitosamente! Los datos han sido registrados en la base del Ministerio.")
    st.balloons()
    if st.button("Finalizar Sesión y Volver al Inicio"):
        st.session_state.clear()
        st.rerun()
    st.stop()

if not st.session_state.iniciado:
    st.info(
        "Estimado(a) rector(a): "
        "Su voz y conocimiento del establecimiento educativo son fundamentales para construir una mirada real y contextualizada sobre la calidad educativa en Colombia. "
        "El diligenciamiento de esta encuesta permitirá contar con información valiosa para reconocer las condiciones en las que se desarrollan los procesos pedagógicos, identificar necesidades prioritarias y fortalecer la toma de decisiones orientadas al mejoramiento educativo. "
        "Cada respuesta contribuirá a visibilizar los avances, retos y capacidades de su institución, y será un insumo clave para continuar promoviendo una educación integral, pertinente, incluyente y de calidad para todos los estudiantes. "
        "Agradecemos profundamente su tiempo, disposición y compromiso con la educación del país."
    )
    df_login = cargar_maestra()

    if df_login.empty:
        st.error("⚠️ No se pudo cargar la base de establecimientos. Verifique el archivo EE 2026.xlsx.")
    else:
        # ── Columnas de referencia ───────────────────────────────────
        col_dep = 'DEPARTAMENTO'          if 'DEPARTAMENTO'          in df_login.columns else df_login.columns[0]
        col_sec = 'SECRETARIA'            if 'SECRETARIA'            in df_login.columns else None
        col_mun = 'MUNICIPIO'             if 'MUNICIPIO'             in df_login.columns else None
        col_nom = 'NOMBRE_ESTABLECIMIENTO' if 'NOMBRE_ESTABLECIMIENTO' in df_login.columns else df_login.columns[1]
        col_dane = 'CODIGO_DANE'          if 'CODIGO_DANE'           in df_login.columns else None

        # ── Filtro 1: Departamento ───────────────────────────────────
        c1, c2 = st.columns(2)
        with c1:
            departamentos = ["Seleccionar..."] + sorted(df_login[col_dep].dropna().unique().tolist())
            depto_sel = st.selectbox("📍 Departamento:", departamentos, key="sel_depto")

        # ── Filtro 2: Secretaría (filtrada por departamento) ─────────
        with c2:
            if depto_sel != "Seleccionar..." and col_sec:
                df_por_depto  = df_login[df_login[col_dep] == depto_sel]
                secretarias   = ["Seleccionar..."] + sorted(df_por_depto[col_sec].dropna().unique().tolist())
                secretaria_sel = st.selectbox("🏛️ Secretaría:", secretarias, key="sel_secretaria")
            else:
                st.selectbox("🏛️ Secretaría:",
                             ["Seleccione primero el Departamento"], disabled=True, key="sel_sec_dis")
                secretaria_sel = "Seleccionar..."

        # ── Filtro 3: Municipio (filtrado por departamento + secretaría) ──
        c3, c4 = st.columns(2)
        with c3:
            if depto_sel != "Seleccionar..." and secretaria_sel != "Seleccionar..." and col_mun:
                df_por_sec    = df_login[
                    (df_login[col_dep] == depto_sel) &
                    (df_login[col_sec] == secretaria_sel)
                ]
                municipios    = ["Seleccionar..."] + sorted(df_por_sec[col_mun].dropna().unique().tolist())
                municipio_sel = st.selectbox("🏙️ Municipio:", municipios, key="sel_municipio")
            else:
                st.selectbox("🏙️ Municipio:",
                             ["Seleccione primero la Secretaría"], disabled=True, key="sel_mun_dis")
                municipio_sel = "Seleccionar..."

        # ── Filtro 4: Colegio (filtrado por municipio, se elige por DANE) ──
        # Mapa de etiqueta visible -> CÓDIGO DANE (único). Diferencia colegios
        # con el mismo nombre en el mismo municipio mostrando el DANE.
        dane_sel = ""
        with c4:
            if (depto_sel != "Seleccionar..." and secretaria_sel != "Seleccionar..."
                    and municipio_sel != "Seleccionar..."):
                df_por_mun  = df_login[
                    (df_login[col_dep] == depto_sel) &
                    (df_login[col_sec] == secretaria_sel) &
                    (df_login[col_mun] == municipio_sel)
                ].copy()

                # Construir etiquetas "NOMBRE — DANE: xxxxx" y mapearlas al DANE
                opciones_map = {}
                for _, r in df_por_mun.iterrows():
                    nom_r  = str(r.get(col_nom, '')).strip()
                    dane_r = str(r.get(col_dane, '')).strip()
                    etiqueta = f"{nom_r}  —  DANE: {dane_r}"
                    opciones_map[etiqueta] = dane_r

                etiquetas   = ["Seleccionar..."] + sorted(opciones_map.keys())
                colegio_sel = st.selectbox("🏫 Establecimiento Educativo:", etiquetas, key="sel_colegio")
                dane_sel    = opciones_map.get(colegio_sel, "")
            else:
                st.selectbox("🏫 Establecimiento Educativo:",
                             ["Seleccione primero el Municipio"], disabled=True, key="sel_col_dis")
                colegio_sel = "Seleccionar..."

        # ── Campo: quien diligencia ──────────────────────────────────
        nombre_diligencia = st.text_input(
            "👤 Nombre de quien diligencia la encuesta:",
            placeholder="Escriba su nombre completo (rector u otro delegado)",
            key="nombre_dilig_login"
        )

        if st.button("🚀 Iniciar Proceso", use_container_width=True):
            if (depto_sel == "Seleccionar..." or secretaria_sel == "Seleccionar..."
                    or municipio_sel == "Seleccionar..." or colegio_sel == "Seleccionar..."):
                st.warning("⚠️ Por favor seleccione el Departamento, la Secretaría, el Municipio y el Establecimiento Educativo.")
            elif not nombre_diligencia.strip():
                st.warning("⚠️ Por favor ingrese el nombre de quien diligencia la encuesta.")
            elif not dane_sel:
                st.error("❌ No se pudo identificar el Código DANE del establecimiento. Verifique la selección.")
            else:
                if es_duplicado_dane(dane_sel):
                    st.error("❌ Esta institución ya cuenta con un registro en el sistema.")
                else:
                    st.session_state.codigo_dane_sel   = dane_sel
                    st.session_state.nombre_diligencia = nombre_diligencia.strip()
                    st.session_state.iniciado          = True
                    st.rerun()

else:
    df_ee = cargar_maestra()
    res   = df_ee[df_ee['CODIGO_DANE'].astype(str).str.strip() == st.session_state.codigo_dane_sel]

    if not res.empty:
        fila = res.iloc[0]

        # --- FASE 1: VERIFICACIÓN ---
        if not st.session_state.encuesta_iniciada:
            st.subheader("Fase 1: Verificación de Datos del Establecimiento")
            bloq = not st.session_state.editando

            col_a, col_b = st.columns(2)
            with col_a:
                nom   = st.text_input("Nombre",                        value=fila.get('NOMBRE_ESTABLECIMIENTO',''), disabled=bloq)
                dan   = st.text_input("Código DANE",                   value=fila.get('CODIGO_DANE',''),            disabled=bloq)
                dep   = st.text_input("Departamento",                  value=fila.get('DEPARTAMENTO',''),           disabled=bloq)
                mun   = st.text_input("Municipio",                     value=fila.get('SECRETARIA',''),             disabled=bloq)
                dan_n = st.text_input("Código DANE Nuevo (Si aplica)", value="",                                    disabled=bloq)
            with col_b:
                rec   = st.text_input("Rector",                        value=fila.get('RECTOR',''),                 disabled=bloq)
                email_f = st.text_input("Correo Electrónico",          value=fila.get('EMAIL',''),                  disabled=bloq)
                dir_e = st.text_input("Dirección",                     value=fila.get('DIRECCION',''),              disabled=bloq)
                bar   = st.text_input("Barrio / Vereda",               value=fila.get('BARRIO_VEREDA',''),          disabled=bloq)
                tel   = st.text_input("Teléfono",                      value=fila.get('TELEFONO',''),               disabled=bloq)
                obs   = st.text_area("Observaciones",                  value="",                                    disabled=bloq, height=55)

            def datos_actuales():
                return {'NOMBRE': nom, 'DANE': dan, 'DEPTO': dep, 'MUNI': mun,
                        'RECTOR': rec, 'EMAIL_VALIDADO': email_f.strip().lower(),
                        'DIR': dir_e, 'BARRIO': bar, 'TEL': tel,
                        'DANE_N': dan_n, 'OBS': obs}

            ca, cb = st.columns(2)
            with ca:
                if not st.session_state.editando:
                    if st.button("✏️ Actualizar"):
                        st.session_state.editando = True; st.rerun()
                else:
                    if st.button("💾 Guardar Cambios"):
                        st.session_state.datos_temp = datos_actuales()
                        st.session_state.editando = False
                        st.session_state.encuesta_iniciada = True; st.rerun()
            with cb:
                if not st.session_state.editando:
                    if st.button("Continuar con Encuesta ➡️"):
                        st.session_state.datos_temp = datos_actuales()
                        st.session_state.encuesta_iniciada = True; st.rerun()

        # --- FASE 2: ENCUESTA ---
        else:
            scroll_al_inicio()
            # Scroll automático al cambiar de bloque
            if st.session_state.hacer_scroll:
                scroll_al_inicio()
                st.session_state.hacer_scroll = False
            st.markdown("### Fase 2: Diligenciamiento de Encuesta")

            paso_actual = st.radio("Módulos:", hojas, index=st.session_state.paso, horizontal=True)
            if hojas.index(paso_actual) != st.session_state.paso:
                st.session_state.paso = hojas.index(paso_actual); st.rerun()

            hoja_act = hojas[st.session_state.paso]
            st.info(f"📍 Usted está en el bloque: **{hoja_act}**")
            obligatorios_hoja_actual = []

            try:
                df_p = pd.read_excel("Encuesta_Calidad.xlsx", sheet_name=hoja_act, skiprows=2)
                for idx, r in df_p.iterrows():
                    p  = str(r.get('Pregunta / campo', r.get('Pregunta', '')))
                    t  = str(r.get('Tipo de respuesta', r.get('Tipo', ''))).lower()
                    o  = str(r.get('Opciones o criterio', r.get('Opciones', '')))
                    v  = str(r.get('Variable', f"{hoja_act}_{idx}")).strip()
                    ob = str(r.get('Obligatorio', '')).strip().lower() == 'sí'

                    if p and p != 'nan':
                        st.write(f"**{p}** {'(*)' if ob else ''}")
                        if ob:
                            obligatorios_hoja_actual.append(v)
                            if v not in st.session_state.oblig:
                                st.session_state.oblig.append(v)

                        if 'lista' in t or 'selección' in t:
                            opts = ["Seleccionar..."] + [x.strip() for x in o.split(';')] if ';' in o else ["Seleccionar...", o]
                            curr = st.session_state.resp_enc.get(v, "Seleccionar...")
                            st.session_state.resp_enc[v] = st.selectbox(
                                "Elegir:", opts, index=opts.index(curr) if curr in opts else 0,
                                key=f"s_{v}", label_visibility="collapsed")
                        elif 'numérico' in t:
                            curr = st.session_state.resp_enc.get(v, 0)
                            st.session_state.resp_enc[v] = st.number_input(
                                "Valor:", min_value=0, step=1, value=int(curr),
                                key=f"n_{v}", label_visibility="collapsed")
                        else:
                            curr = st.session_state.resp_enc.get(v, "")
                            st.session_state.resp_enc[v] = st.text_input(
                                "Escriba:", value=curr,
                                key=f"t_{v}", label_visibility="collapsed")
                        st.write("")
            except Exception as e:
                st.error(f"Error al cargar las preguntas: {e}")

            st.markdown("---")
            c_iz, c_de = st.columns(2)
            with c_iz:
                if st.session_state.paso > 0:
                    if st.button("⬅️ Bloque Anterior"):
                        st.session_state.paso -= 1
                        st.session_state.hacer_scroll = True
                        st.rerun()
            with c_de:
                if st.session_state.paso < len(hojas) - 1:
                    if st.button("Siguiente Bloque ➡️"):
                        faltan = [c for c in obligatorios_hoja_actual
                                  if not st.session_state.resp_enc.get(c)
                                  or str(st.session_state.resp_enc.get(c)).strip() in ["","Seleccionar..."]]
                        if faltan:
                            st.error("⚠️ Por favor responda todas las preguntas obligatorias (*) de este bloque.")
                        else:
                            st.session_state.paso += 1
                            st.session_state.hacer_scroll = True
                            st.rerun()
                else:
                    if st.button("💾 FINALIZAR Y ENVIAR ENCUESTA"):
                        faltan_tot = [c for c in st.session_state.oblig
                                      if not st.session_state.resp_enc.get(c)
                                      or str(st.session_state.resp_enc.get(c)).strip() in ["","Seleccionar..."]]
                        if faltan_tot:
                            st.error("⚠️ Faltan preguntas obligatorias (*) en otros bloques.")
                        else:
                            fecha_f = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            final = {
                                'FECHA':         fecha_f,
                                'DILIGENCIADOR': st.session_state.nombre_diligencia,
                                **st.session_state.datos_temp,
                                **st.session_state.resp_enc
                            }
                            with st.spinner("Guardando encuesta..."):
                                ok, err = guardar_encuesta(final)
                            if ok:
                                if err:
                                    st.warning(err)
                                st.session_state.finalizado = True; st.rerun()
                            else:
                                st.error(f"❌ Error al guardar: {err}.")
    else:
        st.error("Correo no registrado. Por favor verifique.")
        if st.button("Reintentar"):
            st.session_state.clear(); st.rerun()

# =====================================================================
# ACCESO AL PANEL ADMIN
# =====================================================================
st.markdown("<br><br><br>", unsafe_allow_html=True)
with st.expander("🔐 Acceso Administrador"):
    pw = st.text_input("Clave:", type="password", key="pw_admin")
    if pw == "AdminMEN2026":
        if st.button("📊 Abrir Panel de Administración Completo", use_container_width=True):
            st.session_state.panel_admin = True; st.rerun()
    elif pw != "":
        st.error("Acceso denegado.")
